import argparse
import json
import logging
import re
import sys

import requests
from openpyxl import load_workbook

from utils import open_xls_as_xlsx

API_DOMAIN_URL = "https://api.madkudu.com/v1/companies"
API_PERSON_URL = "https://api.madkudu.com/v1/persons"

logger = logging.getLogger('bulk_score')
logger.addHandler(logging.StreamHandler(sys.stdout))
logger.setLevel(logging.DEBUG)


def run_xls(filename: str, api_key: str, score_type: str, column_idx: int):
    print("Welcome to the bulk persons searcher! Wait for the xlsx to load.")
    if re.search('\.xlsx$', filename):
        workbook = load_workbook(filename=filename, keep_vba=False)
        result_filename = filename.replace(".xlsx", ".csv")
    elif re.search('\.xls$', filename):
        workbook = open_xls_as_xlsx(filename)
        result_filename = filename.replace(".xls", ".csv")
    else:
        print("Unsupported file format!")
        exit(1)

    sheet = workbook.active
    regex = re.compile('(?:@)?(?P<tld>[\w\-]+\.\w+)')

    domains_scored = {}
    emails_scored = {}

    print("File loaded. Results will be saved to results/{}.".format(result_filename))
    with open("results/" + result_filename, "a+") as result:
        result.seek(0)
        start = sum(1 for line in result)
        skip_empty = 2
        for row in sheet['A2:B256']:
            if not row[0].value:
                skip_empty += 1
            else:
                break
        try:
            rows = sheet.max_row
            for line in range(start + skip_empty, rows):
                person = {}
                if line % 100 == 0:
                    print("Currently at {}%".format(line / (rows * 1.) * 100.))
                person['email'] = sheet['{}{}'.format(column_idx, line)].value
                if not person['email']:
                    continue

                search = regex.search(person["email"])
                print("scoring: " + person["email"])
                if not search:
                    continue

                if score_type == 'domain':
                    domain = search.group('tld')

                    if domain not in domains_scored:
                        params = {"domain": domain}

                        resp = requests.get(API_DOMAIN_URL, auth=(api_key, ''), params=params)
                        domains_scored[domain] = resp.json()['properties']['customer_fit']

                    customer_fit = domains_scored[domain]
                    result.write(
                        "{},{},{}\n".format(domain, customer_fit['segment'], customer_fit['score'],'"' + format_signals(customer_fit.get('top_signals', '')) + '"')
                    )
                if score_type == 'email':
                    email = person["email"]
                    if email not in emails_scored:
                        params = {"email": email}

                        resp = requests.get(API_PERSON_URL, auth=(api_key, ''), params=params)
                        customer_fit = resp.json()['properties']['customer_fit']
                        emails_scored[email] = resp.json()['properties']['customer_fit']   
                    customer_fit = emails_scored[email]                     
                    result.write(
                        "{},{},{},{}\n".format(email, customer_fit['segment'], customer_fit['score'], '"' + format_signals(customer_fit.get('top_signals', '')) + '"')
                    )        
        except Exception:
            result.flush()
            logger.exception("Exception met. Relaunch to resume!\n")
            exit(1)
        exit(0)

def format_signals(signals: str): 
    return " ".join([format_signal(signal) for signal in signals])

def format_signal(signal: str):
    if not signal:
        return ""
    elif signal["type"] == "positive":
        if signal["value"]:
            return str(json.dumps(str(signal["name"])), 'utf-8').replace('"', '')
        return str('↗ ' + json.dumps(signal["name"]), 'utf-8').replace('"', '')
    elif signal["type"] == "negative":
        if signal["value"]:
            return str('✖ ' + json.dumps(signal["name"]) + ' ' + json.dumps(signal["value"]), 'utf-8').replace('"', '')
        return str('✖ ' + json.dumps(signal["name"]), 'utf-8').replace('"', '')
    return ""

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Sends bulk persons to be scored.')
    parser.add_argument("--filename", help="xlsx file containing all the persons to score", required=True)
    parser.add_argument("--api_key", help="api key", required=True)
    parser.add_argument("--score_type", help="which score to use: either by domain or by personal email", required=True, choices=['domain', 'email'])
    parser.add_argument("--column_idx", help="domain/email column idx (i.e: BQ)", required=True)

    run_xls(**vars(parser.parse_args()))
